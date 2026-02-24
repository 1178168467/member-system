# 第一步：导入所有依赖
import math
import sqlite3
import os
import datetime
import sys
from flask import Flask, render_template, request, jsonify, session, redirect, url_for, g, send_file
import io

# 第二步：创建Flask应用（必须在路由前）
app = Flask(__name__)
app.secret_key = 'member_system_2026_secure_key_' + os.urandom(16).hex()
app.config['TEMPLATES_AUTO_RELOAD'] = True
app.config['JSON_AS_ASCII'] = False  # 解决中文乱码

# ---------------------- 数据库配置（适配Render /tmp目录） ----------------------
DB_PATH = '/tmp/member_system.db'  # Render唯一可写目录

def db_connect():
    """数据库连接（强制用/tmp目录）"""
    if 'db' not in g:
        try:
            g.db = sqlite3.connect(DB_PATH)
            g.db.row_factory = sqlite3.Row
            g.db.execute('PRAGMA foreign_keys = ON')
            g.db.execute('PRAGMA journal_mode = WAL')  # 提升稳定性
        except Exception as e:
            print(f"数据库连接失败：{str(e)}", file=sys.stderr)
            raise
    return g.db

def db_close(e=None):
    """关闭数据库连接"""
    db = g.pop('db', None)
    if db is not None:
        try:
            db.close()
        except:
            pass

def db_query(sql, params=()):
    """数据库查询"""
    try:
        conn = db_connect()
        cursor = conn.cursor()
        cursor.execute(sql, params)
        result = cursor.fetchall()
        return [dict(row) for row in result]
    except Exception as e:
        print(f"查询错误：{str(e)} | SQL: {sql} | 参数: {params}", file=sys.stderr)
        return []

def db_execute(sql, params=()):
    """数据库执行"""
    try:
        conn = db_connect()
        cursor = conn.cursor()
        cursor.execute(sql, params)
        conn.commit()
        return cursor.rowcount
    except Exception as e:
        print(f"执行错误：{str(e)} | SQL: {sql} | 参数: {params}", file=sys.stderr)
        return 0

# ---------------------- 初始化数据库（启动必执行） ----------------------
def init_db():
    """强制初始化所有表"""
    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        cursor.execute('PRAGMA foreign_keys = ON')

        # 1. 创建设置表
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS settings (
                id INTEGER PRIMARY KEY DEFAULT 1,
                shop_name TEXT DEFAULT '',
                shop_phone TEXT DEFAULT '',
                shop_address TEXT DEFAULT '',
                point_rate INTEGER DEFAULT 1,
                level_up_points INTEGER DEFAULT 100,
                level_up_gold_points INTEGER DEFAULT 1000,
                print_receipt INTEGER DEFAULT 1
            )
        ''')

        # 2. 创建操作员表
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS admins (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT UNIQUE NOT NULL,
                password TEXT NOT NULL,
                role TEXT DEFAULT 'operator'
            )
        ''')

        # 3. 创建权限表
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS permissions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                perm_name TEXT UNIQUE NOT NULL,
                perm_desc TEXT DEFAULT ''
            )
        ''')

        # 4. 创建操作员权限关联表
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS admin_permissions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                admin_id INTEGER NOT NULL,
                perm_id INTEGER NOT NULL,
                FOREIGN KEY(admin_id) REFERENCES admins(id) ON DELETE CASCADE,
                FOREIGN KEY(perm_id) REFERENCES permissions(id) ON DELETE CASCADE,
                UNIQUE(admin_id,perm_id)
            )
        ''')

        # 5. 初始化权限数据
        perms = [
            ('member_manage','会员管理'),
            ('recharge_manage','充值管理'),
            ('consume_manage','消费收银'),
            ('points_manage','积分管理'),
            ('report_view','查看报表'),
            ('report_export','导出Excel'),
            ('system_setting','系统设置')
        ]
        for pn,pd in perms:
            cursor.execute('INSERT OR IGNORE INTO permissions(perm_name,perm_desc) VALUES(?,?)',(pn,pd))

        # 6. 创建会员表
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS members (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                card_no TEXT UNIQUE NOT NULL,
                name TEXT NOT NULL,
                phone TEXT UNIQUE NOT NULL,
                level TEXT DEFAULT '普通会员',
                balance REAL DEFAULT 0.0,
                points INTEGER DEFAULT 0,
                create_time DATETIME DEFAULT CURRENT_TIMESTAMP
            )
        ''')

        # 7. 创建消费记录表
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS consume_records (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                member_id INTEGER NOT NULL,
                amount REAL NOT NULL,
                pay_type TEXT NOT NULL,
                remark TEXT DEFAULT '',
                points INTEGER DEFAULT 0,
                create_time DATETIME DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY(member_id) REFERENCES members(id) ON DELETE CASCADE
            )
        ''')

        # 8. 创建充值记录表
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS recharge_records (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                member_id INTEGER NOT NULL,
                amount REAL NOT NULL,
                pay_type TEXT NOT NULL,
                remark TEXT DEFAULT '',
                create_time DATETIME DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY(member_id) REFERENCES members(id) ON DELETE CASCADE
            )
        ''')

        # 9. 初始化管理员账号（admin/admin123）
        cursor.execute('SELECT * FROM admins WHERE username=?',('admin',))
        if not cursor.fetchone():
            cursor.execute('INSERT INTO admins(username,password,role) VALUES(?,?,?)',
                        ('admin','admin123','admin'))

        # 10. 给管理员分配所有权限
        cursor.execute('SELECT id FROM admins WHERE username=?',('admin',))
        admin_row = cursor.fetchone()
        if admin_row:
            aid = admin_row[0]
            allp = db_query('SELECT id FROM permissions')
            for p in allp:
                cursor.execute('INSERT OR IGNORE INTO admin_permissions(admin_id,perm_id) VALUES(?,?)',
                            (aid,p['id']))

        conn.commit()
        conn.close()
        print("✅ 数据库初始化成功", file=sys.stderr)
    except Exception as e:
        print(f"❌ 数据库初始化失败：{str(e)}", file=sys.stderr)
        raise

# 强制重置管理员密码（确保能登录）
def force_reset_admin_pwd():
    try:
        init_db()
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        cursor.execute('UPDATE admins SET password=? WHERE username=?',('admin123','admin'))
        if cursor.rowcount == 0:
            cursor.execute('INSERT INTO admins(username,password,role) VALUES(?,?,?)',
                        ('admin','admin123','admin'))
        conn.commit()
        conn.close()
        print("✅ 管理员密码重置成功", file=sys.stderr)
    except Exception as e:
        print(f"❌ 重置管理员密码失败：{str(e)}", file=sys.stderr)

# 启动时立即初始化数据库
force_reset_admin_pwd()

# ---------------------- 装饰器 ----------------------
def login_required(f):
    """登录验证"""
    def wrapper(*args, **kwargs):
        if 'username' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    wrapper.__name__ = f.__name__
    return wrapper

def permission_required(perm_name):
    """权限验证"""
    def decorator(f):
        def wrapper(*args, **kwargs):
            if 'username' not in session:
                return redirect(url_for('login'))
            if session.get('role') == 'admin':
                return f(*args, **kwargs)
            admin = db_query('SELECT id FROM admins WHERE username=?', (session['username'],))
            if not admin:
                return jsonify({'success':False,'msg':'无权限'})
            aid = admin[0]['id']
            has = db_query('''
                SELECT ap.id FROM admin_permissions ap
                JOIN permissions p ON ap.perm_id=p.id
                WHERE ap.admin_id=? AND p.perm_name=?
            ''',(aid,perm_name))
            if not has:
                return jsonify({'success':False,'msg':'无此操作权限，请联系管理员'})
            return f(*args,**kwargs)
        wrapper.__name__ = f.__name__
        return wrapper
    return decorator

# ---------------------- 会员管理（核心修复） ----------------------
@app.route('/member/add', methods=['POST'])
@login_required
@permission_required('member_manage')
def add_member():
    try:
        # 1. 获取并校验参数
        card_no = request.form.get('card_no','').strip()
        name = request.form.get('name','').strip()
        phone = request.form.get('phone','').strip()
        
        # 校验逻辑
        error_msg = ""
        if not card_no:
            error_msg = "卡号不能为空！"
        elif not name:
            error_msg = "姓名不能为空！"
        elif not phone:
            error_msg = "手机号不能为空！"
        elif not phone.isdigit() or len(phone) != 11:
            error_msg = "手机号必须是11位数字！"
        
        if error_msg:
            return jsonify({'success':False,'msg':error_msg})
        
        # 2. 检查唯一性
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        
        # 检查手机号
        cursor.execute('SELECT id FROM members WHERE phone=?', (phone,))
        if cursor.fetchone():
            conn.close()
            return jsonify({'success':False,'msg':'手机号已存在！'})
        
        # 检查卡号
        cursor.execute('SELECT id FROM members WHERE card_no=?', (card_no,))
        if cursor.fetchone():
            conn.close()
            return jsonify({'success':False,'msg':'卡号已存在！'})
        
        # 3. 插入数据
        insert_sql = '''
            INSERT INTO members (card_no, name, phone, level, balance, points, create_time)
            VALUES (?, ?, ?, '普通会员', 0.0, 0, CURRENT_TIMESTAMP)
        '''
        cursor.execute(insert_sql, (card_no, name, phone))
        conn.commit()
        row_count = cursor.rowcount
        conn.close()
        
        # 4. 返回结果
        if row_count == 1:
            print(f"✅ 会员添加成功：{card_no} | {name} | {phone}", file=sys.stderr)
            return jsonify({'success':True,'msg':'会员添加成功！'})
        else:
            print(f"❌ 会员添加失败：行数={row_count}", file=sys.stderr)
            return jsonify({'success':False,'msg':'添加失败：数据库未写入数据！'})
            
    except sqlite3.Error as e:
        print(f"❌ 数据库错误：{str(e)}", file=sys.stderr)
        return jsonify({'success':False,'msg':f'数据库错误：{str(e)}'})
    except Exception as e:
        print(f"❌ 系统错误：{str(e)}", file=sys.stderr)
        return jsonify({'success':False,'msg':f'添加失败：{str(e)}'})

@app.route('/member/delete/<int:member_id>',methods=['POST'])
@login_required
@permission_required('member_manage')
def delete_member(member_id):
    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        cursor.execute('DELETE FROM members WHERE id=?', (member_id,))
        conn.commit()
        conn.close()
        return jsonify({'success':True,'msg':'删除成功'})
    except Exception as e:
        print(f"❌ 删除会员错误：{str(e)}", file=sys.stderr)
        return jsonify({'success':False,'msg':'删除失败：'+str(e)})

@app.route('/member')
@login_required
@permission_required('member_manage')
def member():
    """会员列表页面"""
    try:
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM members ORDER BY create_time DESC')
        members = [dict(row) for row in cursor.fetchall()]
        conn.close()
        return render_template('member.html',members=members,session=session)
    except Exception as e:
        print(f"❌ 加载会员列表失败：{str(e)}", file=sys.stderr)
        return jsonify({'success':False,'msg':'加载失败：'+str(e)}), 500

# ---------------------- 其他核心路由（完整保留） ----------------------
@app.route('/member/test')
@login_required
def test_member_list():
    """测试接口：查看所有会员"""
    try:
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM members')
        members = [dict(row) for row in cursor.fetchall()]
        conn.close()
        return jsonify({'total':len(members),'members':members})
    except Exception as e:
        return jsonify({'success':False,'msg':str(e)}), 500

@app.route('/recharge')
@login_required
@permission_required('recharge_manage')
def recharge():
    return render_template('recharge.html',session=session)

@app.route('/recharge/search',methods=['POST'])
@login_required
def recharge_search_member():
    kw = request.form.get('keyword','').strip()
    if not kw:
        return jsonify({'success':False,'msg':'请输入'})
    m = db_query('SELECT * FROM members WHERE phone=? OR card_no=?',(kw,kw))
    if not m:
        return jsonify({'success':False,'msg':'未找到'})
    return jsonify({'success':True,'member':m[0]})

@app.route('/recharge/submit',methods=['POST'])
@login_required
@permission_required('recharge_manage')
def submit_recharge():
    try:
        mid = request.form.get('member_id')
        amt = float(request.form.get('amount',0))
        pt = request.form.get('pay_type','无')
        rm = request.form.get('remark','')
        if not mid or amt<=0:
            return jsonify({'success':False,'msg':'金额必须>0'})
        db_execute('INSERT INTO recharge_records(member_id,amount,pay_type,remark) VALUES(?,?,?,?)',
                  (mid,amt,pt,rm))
        db_execute('UPDATE members SET balance=balance+? WHERE id=?',(amt,mid))
        return jsonify({'success':True,'msg':f'充值成功+{amt}元'})
    except Exception as e:
        return jsonify({'success':False,'msg':f'失败：{str(e)}'})

@app.route('/points')
@login_required
@permission_required('points_manage')
def points():
    members = db_query('SELECT id,card_no,name,phone,points,level FROM members ORDER BY points DESC')
    return render_template('points.html',members=members,session=session)

@app.route('/points/adjust',methods=['POST'])
@login_required
@permission_required('points_manage')
def adjust_points():
    try:
        mid = request.form.get('member_id')
        p = int(request.form.get('points',0))
        rm = request.form.get('remark','')
        if not mid:
            return jsonify({'success':False,'msg':'请选择会员'})
        db_execute('UPDATE members SET points=points+? WHERE id=?',(p,mid))
        s = db_query('SELECT level_up_points,level_up_gold_points FROM settings WHERE id=1')
        lu = s[0]['level_up_points'] if s else 100
        lug = s[0]['level_up_gold_points'] if s else 1000
        m = db_query('SELECT points FROM members WHERE id=?',(mid,))
        if m:
            cp = m[0]['points']
            if cp >= lug:
                db_execute('UPDATE members SET level=? WHERE id=?',('金卡会员',mid))
            elif cp >= lu:
                db_execute('UPDATE members SET level=? WHERE id=?',('银卡会员',mid))
            else:
                db_execute('UPDATE members SET level=? WHERE id=?',('普通会员',mid))
        return jsonify({'success':True,'msg':'积分调整成功'})
    except:
        return jsonify({'success':False,'msg':'失败'})

@app.route('/')
@login_required
def index():
    total = len(db_query('SELECT id FROM members'))
    today = datetime.date.today().strftime('%Y-%m-%d')
    tc = db_query('SELECT SUM(amount) as t FROM consume_records WHERE DATE(create_time)=?',(today,))
    tr = db_query('SELECT SUM(amount) as t FROM recharge_records WHERE DATE(create_time)=?',(today,))
    ac = db_query('SELECT SUM(amount) as t FROM consume_records')
    ar = db_query('SELECT SUM(amount) as t FROM recharge_records')
    ap = db_query('SELECT SUM(points) as t FROM members')
    return render_template('index.html',
        username=session.get('username'),
        total_members=total,
        today_consume=tc[0]['t'] if tc and tc[0]['t'] else 0,
        today_recharge=tr[0]['t'] if tr and tr[0]['t'] else 0,
        total_consume=ac[0]['t'] if ac and ac[0]['t'] else 0,
        total_recharge=ar[0]['t'] if ar and ar[0]['t'] else 0,
        total_points=ap[0]['t'] if ap and ap[0]['t'] else 0)

@app.route('/login',methods=['GET','POST'])
def login():
    if request.method=='POST':
        u = request.form.get('username','').strip()
        p = request.form.get('password','').strip()
        if not u or not p:
            return jsonify({'success':False,'msg':'不能为空'})
        a = db_query('SELECT * FROM admins WHERE username=? AND password=?',(u,p))
        if a:
            session['username']=u
            session['role']=a[0]['role']
            return jsonify({'success':True,'msg':'成功','redirect':url_for('index')})
        return jsonify({'success':False,'msg':'账号密码错误'})
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/setting')
@login_required
@permission_required('system_setting')
def setting():
    s = db_query('SELECT * FROM settings WHERE id=1')
    return render_template('setting.html',setting=s[0] if s else None)

@app.route('/setting/save',methods=['POST'])
@login_required
def save_setting():
    try:
        sn = request.form.get('shop_name','')
        sp = request.form.get('shop_phone','')
        sa = request.form.get('shop_address','')
        pr = int(request.form.get('point_rate',1))
        lup = int(request.form.get('level_up_points',100))
        lugp = int(request.form.get('level_up_gold_points',1000))
        prt = int(request.form.get('print_receipt',1))
        if pr<1:
            return jsonify({'success':False,'msg':'比例≥1'})
        if db_query('SELECT id FROM settings WHERE id=1'):
            db_execute('UPDATE settings SET shop_name=?,shop_phone=?,shop_address=?,point_rate=?,level_up_points=?,level_up_gold_points=?,print_receipt=? WHERE id=1',
                      (sn,sp,sa,pr,lup,lugp,prt))
        else:
            db_execute('INSERT INTO settings(id,shop_name,shop_phone,shop_address,point_rate,level_up_points,level_up_gold_points,print_receipt) VALUES(1,?,?,?,?,?,?,?)',
                      (sn,sp,sa,pr,lup,lugp,prt))
        return jsonify({'success':True,'msg':'保存成功'})
    except Exception as e:
        return jsonify({'success':False,'msg':f'失败：{str(e)}'})

@app.route('/setting/get_point_rate')
@login_required
def get_point_rate():
    s = db_query('SELECT point_rate FROM settings WHERE id=1')
    return jsonify({'point_rate': s[0]['point_rate'] if s else 1})

@app.route('/setting/perm_list')
@login_required
def perm_list():
    admins = db_query('SELECT id,username,role FROM admins')
    perms = db_query('SELECT id,perm_name,perm_desc FROM permissions')
    aps = db_query('''
        SELECT ap.admin_id,p.perm_name FROM admin_permissions ap
        JOIN permissions p ON ap.perm_id=p.id
    ''')
    d = {}
    for x in aps:
        if x['admin_id'] not in d:
            d[x['admin_id']]=[]
        d[x['admin_id']].append(x['perm_name'])
    return render_template('perm_list.html',admins=admins,perms=perms,perm_dict=d)

@app.route('/setting/assign_perm',methods=['POST'])
@login_required
def assign_perm():
    if session.get('role')!='admin':
        return jsonify({'success':False,'msg':'仅管理员'})
    aid = request.form.get('admin_id')
    pns = request.form.getlist('perm_names[]')
    db_execute('DELETE FROM admin_permissions WHERE admin_id=?',(aid,))
    for name in pns:
        p = db_query('SELECT id FROM permissions WHERE perm_name=?',(name,))
        if p:
            db_execute('INSERT INTO admin_permissions(admin_id,perm_id) VALUES(?,?)',(aid,p[0]['id']))
    return jsonify({'success':True,'msg':'权限保存成功'})

@app.route('/consume')
@login_required
@permission_required('consume_manage')
def consume():
    return render_template('consume.html')

@app.route('/consume/search',methods=['POST'])
@login_required
def search_member():
    kw = request.form.get('keyword','').strip()
    if not kw:
        return jsonify({'success':False,'msg':'请输入'})
    m = db_query('SELECT * FROM members WHERE phone=? OR card_no=?',(kw,kw))
    if not m:
        return jsonify({'success':False,'msg':'未找到'})
    return jsonify({'success':True,'member':m[0]})

@app.route('/consume/submit',methods=['POST'])
@login_required
def submit_consume():
    try:
        mid = request.form.get('member_id')
        amt = float(request.form.get('amount',0))
        pt = request.form.get('pay_type','无')
        rm = request.form.get('remark','')
        if not mid or amt<=0:
            return jsonify({'success':False,'msg':'金额>0'})

        s = db_query('SELECT point_rate,level_up_points,level_up_gold_points FROM settings WHERE id=1')
        pr = s[0]['point_rate'] if (s and s[0]['point_rate']) else 1
        lu = s[0]['level_up_points'] if (s and s[0]['level_up_points']) else 100
        lug = s[0]['level_up_gold_points'] if (s and s[0]['level_up_gold_points']) else 1000

        add_p = math.floor(amt / pr)
        db_execute('INSERT INTO consume_records(member_id,amount,pay_type,remark,points) VALUES(?,?,?,?,?)',
                  (mid,amt,pt,rm,add_p))
        db_execute('UPDATE members SET points=points+? WHERE id=?',(add_p,mid))

        m = db_query('SELECT points FROM members WHERE id=?',(mid,))
        if m:
            cp = m[0]['points']
            if cp >= lug:
                lv = '金卡会员'
            elif cp >= lu:
                lv = '银卡会员'
            else:
                lv = '普通会员'
            db_execute('UPDATE members SET level=? WHERE id=?',(lv,mid))

        return jsonify({'success':True,'msg':f'成功！按{pr}元1积分，获得{add_p}积分'})
    except Exception as e:
        return jsonify({'success':False,'msg':f'失败：{str(e)}'})

@app.route('/report')
@login_required
@permission_required('report_view')
def report():
    today = datetime.date.today().strftime('%Y-%m-%d')
    month = datetime.date.today().strftime('%Y-%m')
    year = datetime.date.today().strftime('%Y')

    tc = db_query('SELECT SUM(amount) as t FROM consume_records WHERE DATE(create_time)=?',(today,))
    mc = db_query('SELECT SUM(amount) as t FROM consume_records WHERE strftime("%Y-%m",create_time)=?',(month,))
    yc = db_query('SELECT SUM(amount) as t FROM consume_records WHERE strftime("%Y",create_time)=?',(year,))
    tr = db_query('SELECT SUM(amount) as t FROM recharge_records WHERE DATE(create_time)=?',(today,))
    mr = db_query('SELECT SUM(amount) as t FROM recharge_records WHERE strftime("%Y-%m",create_time)=?',(month,))
    yr = db_query('SELECT SUM(amount) as t FROM recharge_records WHERE strftime("%Y",create_time)=?',(year,))

    ln = db_query('SELECT COUNT(*) as c FROM members WHERE level=?',('普通会员',))
    ls = db_query('SELECT COUNT(*) as c FROM members WHERE level=?',('银卡会员',))
    lg = db_query('SELECT COUNT(*) as c FROM members WHERE level=?',('金卡会员',))

    return render_template('report.html',
        today_consume=tc[0]['t'] if tc and tc[0]['t'] else 0,
        month_consume=mc[0]['t'] if mc and mc[0]['t'] else 0,
        year_consume=yc[0]['t'] if yc and yc[0]['t'] else 0,
        today_recharge=tr[0]['t'] if tr and tr[0]['t'] else 0,
        month_recharge=mr[0]['t'] if mr and mr[0]['t'] else 0,
        year_recharge=yr[0]['t'] if yr and yr[0]['t'] else 0,
        level_normal=ln[0]['c'] if ln else 0,
        level_silver=ls[0]['c'] if ls else 0,
        level_gold=lg[0]['c'] if lg else 0,
        session=session)

@app.route('/report/export')
@login_required
@permission_required('report_export')
def export_report():
    try:
        from openpyxl import Workbook
        wb = Workbook()
        ws1 = wb.active
        ws1.title = '会员列表'
        ws1.append(['卡号','姓名','手机号','等级','余额','积分'])
        ms = db_query('SELECT card_no,name,phone,level,balance,points FROM members')
        for m in ms:
            ws1.append([m['card_no'],m['name'],m['phone'],m['level'],m['balance'],m['points']])

        ws2 = wb.create_sheet('充值记录')
        ws2.append(['会员卡号','金额','时间'])
        rs = db_query('''
            SELECT m.card_no,r.amount,r.create_time
            FROM recharge_records r
            JOIN members m ON r.member_id=m.id
            ORDER BY r.create_time DESC
        ''')
        for r in rs:
            ws2.append([r['card_no'],r['amount'],r['create_time']])

        ws3 = wb.create_sheet('消费记录')
        ws3.append(['卡号','金额','积分','时间'])
        cs = db_query('''
            SELECT m.card_no,c.amount,c.points,c.create_time
            FROM consume_records c
            JOIN members m ON c.member_id=m.id
            ORDER BY c.create_time DESC
        ''')
        for c in cs:
            ws3.append([c['card_no'],c['amount'],c['points'],c['create_time']])

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        fn = f'会员数据_{datetime.datetime.now().strftime("%Y%m%d%H%M%S")}.xlsx'
        return send_file(buf,download_name=fn,as_attachment=True)
    except Exception as e:
        print(f"❌ 导出Excel错误：{str(e)}", file=sys.stderr)
        return jsonify({'success':False,'msg':f'导出失败：{str(e)}'})

# ---------------------- 启动配置 ----------------------
app.teardown_appcontext(db_close)

# Render 部署配置
if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
